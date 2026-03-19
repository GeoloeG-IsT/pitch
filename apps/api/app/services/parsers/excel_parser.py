"""Excel ingestion: openpyxl + gpt-4o-mini summary + OpenAIEmbedding.

Each worksheet becomes one chunk with:
- content: LLM-generated natural language summary
- metadata.raw_data: the actual row data
- metadata.sheet_name: worksheet name
- chunk_type: "table"
"""

from __future__ import annotations

import io

import tiktoken
from llama_index.embeddings.openai import OpenAIEmbedding
from openai import OpenAI
from openpyxl import load_workbook

_encoding = tiktoken.encoding_for_model("text-embedding-3-small")
_embed_model: OpenAIEmbedding | None = None


def _get_embed_model() -> OpenAIEmbedding:
    """Lazy-init so the model picks up OPENAI_API_KEY set at startup."""
    global _embed_model
    if _embed_model is None:
        _embed_model = OpenAIEmbedding(model="text-embedding-3-small")
    return _embed_model


def _rows_to_text(rows: list[list[str]]) -> str:
    """Format rows as tab-separated text with row numbers."""
    lines: list[str] = []
    for i, row in enumerate(rows, start=1):
        lines.append(f"{i}\t{'\t'.join(row)}")
    return "\n".join(lines)


def _generate_summary(sheet_name: str, table_text: str) -> str:
    """Generate a natural language summary of a spreadsheet sheet via gpt-4o-mini."""
    client = OpenAI()
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": (
                    "Summarize this financial spreadsheet data in clear natural language. "
                    "Include all key numbers, metrics, and relationships. "
                    "Be comprehensive -- this summary will be used for semantic search."
                ),
            },
            {
                "role": "user",
                "content": f"Sheet: {sheet_name}\n\n{table_text}",
            },
        ],
        max_tokens=1000,
    )
    return response.choices[0].message.content


async def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse an Excel file into chunk dicts, one per non-empty worksheet.

    Each chunk contains an LLM-generated summary as content and the raw
    row data in metadata. Embeddings are generated via OpenAI text-embedding-3-small.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    chunks: list[dict] = []

    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        # Skip empty sheets
        if not any(any(cell for cell in row) for row in rows):
            continue

        table_text = _rows_to_text(rows)
        summary = _generate_summary(sheet_name, table_text)

        embedding = await _get_embed_model().aget_text_embedding(summary)
        token_count = len(_encoding.encode(summary))

        chunks.append({
            "content": summary,
            "embedding": embedding,
            "section_number": idx,
            "page_number": None,
            "chunk_type": "table",
            "metadata": {
                "sheet_name": sheet_name,
                "raw_data": rows,
                "row_count": len(rows),
                "col_count": max(len(r) for r in rows) if rows else 0,
            },
            "token_count": token_count,
        })

    wb.close()
    return chunks

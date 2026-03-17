#!/usr/bin/env python3
"""Build financial model Excel workbook with formulas and formatting."""
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

SCRIPT_DIR = pathlib.Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR.parent / "output"
OUTPUT = OUTPUT_DIR / "financial-model.xlsx"

HEADER_FILL = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, color="666666", size=10)
CURRENCY_FORMAT = '"$"#,##0'
CURRENCY_K_FORMAT = '"$"#,##0"K"'
PERCENT_FORMAT = "0%"


def auto_width(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)


def add_title(ws, title, cols=4):
    """Add a merged title row to the sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center")


def style_headers(ws, row, col_count):
    """Apply header styling to a row."""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def build_tam_sam_som(wb):
    """Sheet 1: TAM SAM SOM market sizing."""
    ws = wb.active
    ws.title = "TAM SAM SOM"
    add_title(ws, "Market Sizing: TAM / SAM / SOM", 4)

    headers = ["Segment", "Value ($B)", "Methodology", "Key Assumptions"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_headers(ws, 3, 4)

    data = [
        [
            "Total Addressable Market (TAM)",
            300,
            "Top-down: total global venture capital investment volume (2025)",
            "All VC deal flow globally, all stages, all sectors",
        ],
        [
            "Serviceable Available Market (SAM)",
            15,
            "Top-down filtered: tech vertical, early-stage rounds, target geographies",
            "Tech startups raising Seed through Series B, North America + Europe",
        ],
        [
            "Serviceable Obtainable Market (SOM)",
            0.15,
            "Bottom-up: target 5,000 startups x $30K avg annual spend",
            "First-mover AI pitch platforms, 1% of SAM achievable by Year 3",
        ],
    ]

    for r, row_data in enumerate(data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 2:
                cell.number_format = '"$"#,##0.00"B"'

    ws.cell(row=8, column=1, value="Referenced in Pitch Deck, Slide 8: Market Opportunity").font = NOTE_FONT
    auto_width(ws)


def build_revenue_projections(wb):
    """Sheet 2: Revenue Projections with formulas."""
    ws = wb.create_sheet("Revenue Projections")
    add_title(ws, "3-Year Revenue Projections", 4)

    headers = ["Metric", "Year 1", "Year 2", "Year 3"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_headers(ws, 3, 4)

    # Row 4: Customers
    ws.cell(row=4, column=1, value="Customers")
    ws.cell(row=4, column=2, value=50)
    ws.cell(row=4, column=3, value=250)
    ws.cell(row=4, column=4, value=1000)

    # Row 5: Avg Monthly Subscription
    ws.cell(row=5, column=1, value="Avg Monthly Subscription ($)")
    ws.cell(row=5, column=2, value=500)
    ws.cell(row=5, column=3, value=500)
    ws.cell(row=5, column=4, value=500)
    for c in range(2, 5):
        ws.cell(row=5, column=c).number_format = CURRENCY_FORMAT

    # Row 6: MRR = Customers * Avg Monthly Subscription (FORMULA)
    ws.cell(row=6, column=1, value="Monthly Recurring Revenue (MRR)")
    ws.cell(row=6, column=2, value="=B4*B5")
    ws.cell(row=6, column=3, value="=C4*C5")
    ws.cell(row=6, column=4, value="=D4*D5")
    for c in range(2, 5):
        ws.cell(row=6, column=c).number_format = CURRENCY_FORMAT

    # Row 7: ARR = MRR * 12 (FORMULA)
    ws.cell(row=7, column=1, value="Annual Recurring Revenue (ARR)")
    ws.cell(row=7, column=2, value="=B6*12")
    ws.cell(row=7, column=3, value="=C6*12")
    ws.cell(row=7, column=4, value="=D6*12")
    for c in range(2, 5):
        ws.cell(row=7, column=c).number_format = CURRENCY_FORMAT

    # Row 8: Usage Revenue
    ws.cell(row=8, column=1, value="Usage Revenue (overage)")
    ws.cell(row=8, column=2, value=60000)
    ws.cell(row=8, column=3, value=450000)
    ws.cell(row=8, column=4, value=2400000)
    for c in range(2, 5):
        ws.cell(row=8, column=c).number_format = CURRENCY_FORMAT

    # Row 9: Total Revenue = ARR + Usage Revenue (FORMULA)
    ws.cell(row=9, column=1, value="Total Revenue")
    ws.cell(row=9, column=2, value="=B7+B8")
    ws.cell(row=9, column=3, value="=C7+C8")
    ws.cell(row=9, column=4, value="=D7+D8")
    for c in range(2, 5):
        ws.cell(row=9, column=c).number_format = CURRENCY_FORMAT
        ws.cell(row=9, column=c).font = Font(bold=True)

    # Row 10: Growth Rate YoY (FORMULA)
    ws.cell(row=10, column=1, value="Growth Rate YoY")
    ws.cell(row=10, column=2, value="N/A")
    ws.cell(row=10, column=3, value="=(C9-B9)/B9")
    ws.cell(row=10, column=4, value="=(D9-C9)/C9")
    for c in range(3, 5):
        ws.cell(row=10, column=c).number_format = PERCENT_FORMAT

    ws.cell(row=12, column=1, value="Referenced in Pitch Deck, Slide 9: Business Model").font = NOTE_FONT
    auto_width(ws)


def build_unit_economics(wb):
    """Sheet 3: Unit Economics with formulas."""
    ws = wb.create_sheet("Unit Economics")
    add_title(ws, "Unit Economics Analysis", 4)

    headers = ["Metric", "Year 1", "Year 2", "Year 3"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_headers(ws, 3, 4)

    # Row 4: CAC
    ws.cell(row=4, column=1, value="Customer Acquisition Cost (CAC)")
    ws.cell(row=4, column=2, value=2000)
    ws.cell(row=4, column=3, value=1500)
    ws.cell(row=4, column=4, value=1000)
    for c in range(2, 5):
        ws.cell(row=4, column=c).number_format = CURRENCY_FORMAT

    # Row 5: Average Lifetime (months)
    ws.cell(row=5, column=1, value="Avg Customer Lifetime (months)")
    ws.cell(row=5, column=2, value=36)
    ws.cell(row=5, column=3, value=36)
    ws.cell(row=5, column=4, value=36)

    # Row 6: Monthly Subscription
    ws.cell(row=6, column=1, value="Monthly Subscription ($)")
    ws.cell(row=6, column=2, value=500)
    ws.cell(row=6, column=3, value=500)
    ws.cell(row=6, column=4, value=500)
    for c in range(2, 5):
        ws.cell(row=6, column=c).number_format = CURRENCY_FORMAT

    # Row 7: LTV = Lifetime * Monthly Subscription (FORMULA)
    ws.cell(row=7, column=1, value="Lifetime Value (LTV)")
    ws.cell(row=7, column=2, value="=B5*B6")
    ws.cell(row=7, column=3, value="=C5*C6")
    ws.cell(row=7, column=4, value="=D5*D6")
    for c in range(2, 5):
        ws.cell(row=7, column=c).number_format = CURRENCY_FORMAT

    # Row 8: LTV:CAC Ratio (FORMULA)
    ws.cell(row=8, column=1, value="LTV:CAC Ratio")
    ws.cell(row=8, column=2, value="=B7/B4")
    ws.cell(row=8, column=3, value="=C7/C4")
    ws.cell(row=8, column=4, value="=D7/D4")
    for c in range(2, 5):
        ws.cell(row=8, column=c).number_format = "0.0"

    # Row 9: CAC Payback Period = CAC / Monthly Subscription (FORMULA)
    ws.cell(row=9, column=1, value="CAC Payback Period (months)")
    ws.cell(row=9, column=2, value="=B4/B6")
    ws.cell(row=9, column=3, value="=C4/C6")
    ws.cell(row=9, column=4, value="=D4/D6")
    for c in range(2, 5):
        ws.cell(row=9, column=c).number_format = "0.0"

    # Row 10: Gross Margin
    ws.cell(row=10, column=1, value="Gross Margin")
    ws.cell(row=10, column=2, value=0.85)
    ws.cell(row=10, column=3, value=0.85)
    ws.cell(row=10, column=4, value=0.85)
    for c in range(2, 5):
        ws.cell(row=10, column=c).number_format = PERCENT_FORMAT

    ws.cell(row=12, column=1, value="Referenced in Investment Memo, Section: Unit Economics Analysis").font = NOTE_FONT
    auto_width(ws)


def build_burn_rate(wb):
    """Sheet 4: Burn Rate & Runway."""
    ws = wb.create_sheet("Burn Rate & Runway")
    add_title(ws, "Burn Rate & Runway Analysis", 2)

    headers = ["Category", "Monthly ($K)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_headers(ws, 3, 2)

    # Expense rows
    ws.cell(row=4, column=1, value="Engineering (salaries + infra)")
    ws.cell(row=4, column=2, value=40)

    ws.cell(row=5, column=1, value="Go-to-Market")
    ws.cell(row=5, column=2, value=15)

    ws.cell(row=6, column=1, value="Operations")
    ws.cell(row=6, column=2, value=10)

    # Total Monthly Burn (FORMULA)
    ws.cell(row=7, column=1, value="Total Monthly Burn")
    ws.cell(row=7, column=2, value="=SUM(B4:B6)")
    ws.cell(row=7, column=1).font = Font(bold=True)
    ws.cell(row=7, column=2).font = Font(bold=True)

    for c in range(2, 3):
        for r in range(4, 8):
            ws.cell(row=r, column=c).number_format = CURRENCY_K_FORMAT

    # Funding scenarios
    ws.cell(row=9, column=1, value="Funding Scenarios").font = Font(bold=True, size=12)

    ws.cell(row=10, column=1, value="At $1M raise: Runway (months)")
    ws.cell(row=10, column=2, value="=1000/B7")
    ws.cell(row=10, column=2).number_format = "0.0"

    ws.cell(row=11, column=1, value="At $2M raise: Runway (months)")
    ws.cell(row=11, column=2, value="=2000/B7")
    ws.cell(row=11, column=2).number_format = "0.0"

    ws.cell(row=13, column=1, value="Referenced in Pitch Deck, Slide 13: The Ask").font = NOTE_FONT
    auto_width(ws)


def build_assumptions(wb):
    """Sheet 5: Key Assumptions."""
    ws = wb.create_sheet("Assumptions")
    add_title(ws, "Key Model Assumptions", 2)

    headers = ["Assumption", "Value / Source"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_headers(ws, 3, 2)

    assumptions = [
        ("Average contract length", "36 months"),
        ("Monthly churn rate", "3%"),
        ("Annual expansion revenue", "10%"),
        ("Cost of AI inference per query", "$0.003"),
        ("Avg queries per investor session", "15"),
        ("Avg investor views per pitch room per month", "20"),
        ("Embedding model", "OpenAI text-embedding-3-small (1536 dimensions)"),
        ("Target gross margin", "85%"),
        ("SaaS base subscription", "$500/month per pitch room"),
        ("Usage overage pricing", "Per-investor-view and per-query above threshold"),
    ]

    for r, (assumption, value) in enumerate(assumptions, 4):
        ws.cell(row=r, column=1, value=assumption)
        ws.cell(row=r, column=2, value=value)

    note_row = 4 + len(assumptions) + 1
    ws.cell(
        row=note_row,
        column=1,
        value="All figures are illustrative projections for demo purposes. See Investment Memo for market timing rationale.",
    ).font = NOTE_FONT
    auto_width(ws)


def build():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    build_tam_sam_som(wb)
    build_revenue_projections(wb)
    build_unit_economics(wb)
    build_burn_rate(wb)
    build_assumptions(wb)

    wb.save(str(OUTPUT))
    print(f"Built: {OUTPUT}")


if __name__ == "__main__":
    build()
